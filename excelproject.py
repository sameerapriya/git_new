import openpyxl as xl
from openpyxl.chart import BarChart,Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)#accessing the workbook
    sheet = wb['Sheet1'] #accessing the sheet
    #cell = sheet['a1']
    #cell = sheet.cell(1 , 1)
    #print(cell.value) #what is present in that particular cell
    #print(sheet.max_row)#for number of rows

    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3) #accessing elements
        corrected_price = cell.value*0.9 # to apply the formula
        corrected_price_cell = sheet.cell(row,4) #assigning them cells via object
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
    #creating instance of a reference class and selecting the elements
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e8')
    wb.save(filename)

process_workbook("transactions.xlsx")